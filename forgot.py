from tkinter import *
from tkinter import messagebox
import openpyxl
from PIL import ImageTk, Image
from tkinter import ttk
# sh1.cell(row=i,column=j,value=g)


def des(root: Tk):
    root.destroy()


def fp(a: Entry, root: Tk):
    wb = openpyxl.load_workbook("data.xlsx")
    sh1 = wb['student']
    mail = a.get()
    flag = 1
    row = sh1.max_row
    for i in range(1, row+1):
        if sh1.cell(row=i, column=1).value == mail:
            flag =0
            password = sh1.cell(row=i, column=3).value
            user = sh1.cell(row=i, column=2).value
            f = Label(root, text=f"Your Username is :: {user}", font=("bold", 15))
            f.place(x=20, y=200)
            e = Label(root, text=f"Your Password is :: {password}", font=("bold", 15))
            e.place(x=20, y=250)
            break
    if flag == 1:
        e = Label(root, text="Email id is not registered", font=("bold", 15))
        e.place(x=20, y=200)


def forgot():
    root2 = Tk()
    root2.geometry("1200x600+0+0")
    root2.title("Forgot password page")
    # def resize_image(event):
    #     new_width = event.width
    #     new_height = event.height
    #     image = copy_of_image.resize((new_width, new_height))
    #     photo = ImageTk.PhotoImage(image)
    #     label.config(image=photo)
    #     label.image = photo  # avoid garbage collection
    #
    # image = Image.open('images.jpg')
    # copy_of_image = image.copy()
    # photo = ImageTk.PhotoImage(image)
    # label = ttk.Label(root2, image=photo)
    # label.bind('<Configure>', resize_image)
    # label.pack(fill=BOTH, expand=YES)


    title = Label(root2, text="Forgot Password", font=("Impact", 35, "bold"), fg="#d77337")
    title.place(x=90, y=30)

    email = Label(root2, text="Enter Email Id::", font=("bold", 12))
    email.place(x=20, y=100)

    eMail = Entry(root2, font=("times new roman", 15))
    eMail.place(x=150, y=100, height=35, width=350)

    get = Button(root2, text="Get password", font=("bold", 15), bg="pink", command=lambda: fp(eMail, root2))
    get.place(x=100, y=150)

    done = Button(root2, text="Done", font=("bold", 15), bg="pink", command=lambda: des(root2))
    done.place(x=300, y=150)
    root2.mainloop()
# forgot()