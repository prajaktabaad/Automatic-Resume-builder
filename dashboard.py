from tkinter import *
from PIL import Image, ImageTk
from tkinter import messagebox,ttk,filedialog
from fpdf import FPDF
import openpyxl
import PyPDF2
import subprocess
import variable


def getRownum(u: str, row: int):
    wb = openpyxl.load_workbook("data.xlsx")
    sh1 = wb['student']
    m = 0
    for i in range(2, row + 1):
        if u == sh1.cell(row=i, column=2).value:
            m = i
    return m


def browseFiles(u:str):
    wb = openpyxl.load_workbook("data.xlsx")
    sh1 = wb['student']
    row = sh1.max_row
    r = int(getRownum(u, row))

    f = filedialog.askopenfilename(title = "Select a File",
                                          filetypes = (("Image files",
                                                        "*.jpeg*"),
                                                       ("all files",
                                                        "*.*")))
    sh1.cell(row=r, column=11, value=f)
    wb.save("data.xlsx")


def ctpdf(u: str, user_phone: Entry, user_lnk: Entry, user_edu: Text, user_skill: Text, user_pro: Text,
          user_cert: Text):
    wb = openpyxl.load_workbook("data.xlsx")
    sh1 = wb['student']

    upn = user_phone.get()
    ulnk = user_lnk.get()
    uedu = user_edu.get(1.0, "end-1c")
    uskil = user_skill.get(1.0, "end-1c")
    ucert = user_cert.get(1.0, "end-1c")
    uproj = user_pro.get(1.0, "end-1c")

    list1 = [uedu, ulnk, uskil, uproj, ucert, upn]
    flag = 1
    row = sh1.max_row
    col = sh1.max_column

    r = int(getRownum(u, row))
    email = sh1.cell(row=r, column=1).value

    for j in range(5, 11):
        if list1[j - 5] == "":
            continue
        sh1.cell(row=r, column=j, value=list1[j - 5])
    messagebox.showinfo("Status", "Entered Successfully")
    wb.save("data.xlsx")

    flist = []
    # flist = [email, name, edu, ulnk, uskil, uproj, ucert, upn, img]
    for j in range(1, 12):
        if j == 2 or j == 3:
            continue
        flist.append(sh1.cell(row=r, column=j).value)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=25)
    pdf.image(flist[8], x=150, y=10, w=30, h=30, type='jpeg')
    pdf.set_xy(10, 10)
    pdf.set_text_color(0, 0, 255)
    pdf.cell(200, 10, txt=f"{flist[1]}") # name
    pdf.set_font("Arial", size=15)
    pdf.set_xy(12, 17)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(200, 10, txt=f"{flist[0]}") # email
    pdf.set_xy(12, 24)
    pdf.cell(200, 10, txt=f"{flist[7]}")  # phone
    pdf.set_xy(12, 31)
    pdf.cell(200, 10, txt=f"{flist[3]}")  # linkedin

    pdf.line(10, 41, 200, 41)

    pdf.set_xy(10, 45)
    pdf.set_font("Arial", size=20)
    pdf.cell(200, 10, txt="Education:")  # edu1
    pdf.set_xy(15, 52)
    pdf.set_font("Arial", size=15)
    pdf.cell(200, 10, txt=f"{flist[2]}")  # edu2

    pdf.set_xy(10, 95)
    pdf.set_font("Arial", size=20)
    pdf.cell(200, 10, txt="Projects:")  # proj1
    pdf.set_xy(15, 102)
    pdf.set_font("Arial", size=15)
    pdf.cell(200, 10, txt=f"{flist[5]}")  # proj2

    pdf.set_xy(10, 145)
    pdf.set_font("Arial", size=20)
    pdf.cell(200, 10, txt="Skills:")  # skill 1
    pdf.set_xy(15, 152)
    pdf.set_font("Arial", size=15)
    pdf.cell(200, 10, txt=f"{flist[4]}")  # skill 2

    pdf.set_xy(10, 195)
    pdf.set_font("Arial", size=20)
    pdf.cell(200, 10, txt="Certification:")  # cert1
    pdf.set_xy(15, 202)
    pdf.set_font("Arial", size=15)
    pdf.cell(200, 10, txt=f"{flist[6]}")  # cert2

    pdfnamne = u + ".pdf"
    pdf.output(pdfnamne)

    subprocess.Popen([pdfnamne], shell=True)


def dash(uname: str):
    root3 = Tk()
    root3.geometry("1200x600+0+0")
    root3.title("db")
    def resize_image(event):
        new_width = event.width
        new_height = event.height
        image = copy_of_image.resize((new_width, new_height))
        photo = ImageTk.PhotoImage(image)
        label.config(image=photo)
        label.image = photo  # avoid garbage collection

    image = Image.open('images.jpg')
    copy_of_image = image.copy()
    photo = ImageTk.PhotoImage(image)
    label = ttk.Label(root3, image=photo)
    label.bind('<Configure>', resize_image)
    label.pack(fill=BOTH, expand=YES)

    wb = openpyxl.load_workbook("data.xlsx")
    sh1 = wb['student']
    flag = 1
    row = sh1.max_row
    for i in range(2, row + 1):
        if sh1.cell(row=i, column=2).value == uname:
            u = sh1.cell(row=i, column=4).value
            e = Label(root3, text=f"Welcome::{u}::Please enter your details!", font=("bold", 15), bg="white")
            e.place(x=20, y=15)

    phone = Label(root3, text="Phone no. ", font=("Century", 15, "bold"), bg="white", fg="black").place(
        x=50, y=70)
    user_phone = Entry(root3, font=("times new roman", 15), bg="white")
    user_phone.place(x=170, y=80, height=25, width=250)

    lnk = Label(root3, text="LinkedIn ", font=("Century", 15, "bold"), bg="white", fg="black").place(
        x=450, y=70)
    user_lnk = Entry(root3, font=("times new roman", 15), bg="white")
    user_lnk.place(x=570, y=80, height=25, width=250)

    edu = Label(root3, text="Education ", font=("Century", 15, "bold"), bg="white", fg="black").place(
        x=50, y=145)
    user_edu = Text(root3, font=("times new roman", 15), bg="white")
    user_edu.place(x=170, y=125, height=135, width=250)

    skill = Label(root3, text="Skills ", font=("Century", 15, "bold"), bg="white", fg="black").place(
        x=480, y=145)
    user_skill = Text(root3, font=("times new roman", 15), bg="white")
    user_skill.place(x=570, y=125, height=135, width=250)

    pro = Label(root3, text="Projects", font=("Century", 15, "bold"), bg="white", fg="black").place(
        x=50, y=305)
    user_pro = Text(root3, font=("times new roman", 15), bg="white")
    user_pro.place(x=170, y=305, height=135, width=250)

    cert = Label(root3, text="Certification", font=("Century", 14, "bold"), bg="white", fg="black").place(
        x=430, y=305)
    user_cert = Text(root3, font=("times new roman", 15), bg="white")
    user_cert.place(x=570, y=305, height=135, width=250)
    # print(type(user_cert))
    done = Button(root3, text="Download and View", font=("bold", 15), bg="pink", command=lambda: ctpdf(uname,
                                                                                                      user_phone,
                                                                                             user_lnk,
                                                                                          user_edu, user_skill,
                                                                                          user_pro, user_cert))
    done.place(x=570, y=500)
    addimg = Button(root3, text="Upload Photo", font=("bold", 15), bg="white", command=lambda: browseFiles(uname))

    addimg.place(x=170, y=500)
    root3.mainloop()


# dash("ankit")
