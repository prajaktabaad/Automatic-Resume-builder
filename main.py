from tkinter import *
from PIL import Image, ImageTk
from tkinter import messagebox
import openpyxl
from tkinter import ttk

import register as reg
import forgot as fg
import dashboard as db


def fp():
    # to call forgot password
    fg.forgot()


def re():
    # To call register
    reg.register()


def login(txt_user: Entry, txt_userpass: Entry, root: Tk):
    p = txt_userpass.get()
    u = txt_user.get()

    wb = openpyxl.load_workbook("data.xlsx")
    sh1 = wb['student']

    row = sh1.max_row
    col = sh1.max_column
    flag = 1
    if u == "" or p == "":
        messagebox.showerror("Error", "All fields are required", parent=root)
    else:
        for i in range(2, row + 1):
            if u == sh1.cell(row=i, column=2).value and p == sh1.cell(row=i, column=3).value:
                flag = 0
                messagebox.showinfo("Done!!", "Successful Login!!", parent=root)
                root.destroy()
                db.dash(u)
        if flag == 1:
            messagebox.showerror("Error", "Please Enter Correct username and password", parent=root)


def mainpg():
    root = Tk()
    root.title("Login page")
    root.geometry("1200x600+0+0")

    def resize_image(event):
        new_width = event.width
        new_height = event.height
        image = copy_of_image.resize((new_width, new_height))
        photo = ImageTk.PhotoImage(image)
        label.config(image=photo)
        label.image = photo  # avoid garbage collection

    image = Image.open('images.jpg')
    copy_of_image = image.copy()
    photo = ImageTk.PhotoImage(image)
    label = ttk.Label(root, image=photo)
    label.bind('<Configure>', resize_image)
    label.pack(fill=BOTH, expand=YES)

    frame_login = Frame(root, bg="white")
    frame_login.place(x=15, y=27, height=500, width=600)
    # root.config(bg="Lavender")
    # bg = ImageTk.PhotoImage(file="try2.jpg")

    title = Label(frame_login, text="Resume builder", font=("Impact", 35, "bold"), bg="white", fg="#d77337").place(x=90,
                                                                                                               y=30)
    desc = Label(frame_login, text="Login Here", font=("Impact", 20, "bold"), bg="white",
                 fg="#d77337").place(x=90, y=100)

    user_label = Label(frame_login, text="Username : ", font=("Century", 15, "bold"), bg="white", fg="black").place(
        x=90, y=140)

    txt_user = Entry(frame_login, font=("times new roman", 15), bg="white")
    txt_user.place(x=90, y=170, height=35, width=350)

    userpass_label = Label(frame_login, text="Password", font=("Century", 15, "bold"), bg="white", fg="black").place(
        x=90, y=220)
    txt_userpass = Entry(frame_login, font=("times new roman", 15), bg="white", show="*")
    txt_userpass.place(x=90, y=250, height=35, width=350)

    login_btn = Button(frame_login, text="Login", command=lambda: login(txt_user, txt_userpass, root), bg="#d77337",
                       fg="white", font="Calibry 12 bold", ).place(x=90, y=300, height=35, width=160)
    forgot_btn = Button(frame_login, text="Forgot Password?", command=fp, bg="#d77337",
                        fg="white", font="Calibry 12 bold", ).place(x=280, y=300, height=35, width=160)

    label1 = Label(frame_login, text="New registration?", font=("Century", 15, "bold"), bg="white", fg="black").place(
        x=90, y=350)
    register = Button(frame_login, text="Register Here", bg="#d77337", font="Calibry 12 bold", fg="white",
                      command=re).place(x=290, y=350, height=35, width=150)

    root.mainloop()


mainpg()
