from tkinter import *
from tkinter import messagebox
import openpyxl
from tkinter import ttk
from PIL import ImageTk, Image

def des(root: Tk):
    root.destroy()


def reg(a: Entry, b: Entry, c: Entry, d: Entry, root: Tk):
    wb = openpyxl.load_workbook("data.xlsx")
    sh1 = wb['student']

    uname1 = b.get()
    name = d.get()
    pas = c.get()
    mail = a.get()

    list1 = [mail, uname1, pas, name]
    flag = 1
    row = sh1.max_row
    col = sh1.max_column
    if uname1 == "" or pas == "" or name == "" or mail == "":
        messagebox.showinfo("Error", "All fields are required in Register")
    else:
        for i in range(2, row+1):
            if sh1.cell(row=i, column=1).value == mail:
                flag = 0
                messagebox.showinfo("Error", "Mail id already registered!\nPlease enter other")
            elif sh1.cell(row=i, column=1).value == uname1:
                flag = 0
                messagebox.showinfo("Error", "Username already taken!\nPlease enter other")
        if flag == 1:
            for j in range(1, 5):
                sh1.cell(row=row + 1, column=j, value=list1[j - 1])
            messagebox.showinfo("Status", "Registered Successfully")
            wb.save("data.xlsx")
            root.destroy()


def register():
    root1 = Tk()
    root1.geometry("1200x600+0+0")
    root1.title("Register page")

    # def resize_image(event):
    #     new_width = event.width
    #     new_height = event.height
    #     image = copy_of_image.resize((new_width, new_height))
    #     photo = ImageTk.PhotoImage(image)
    #     label.config(image=photo)
    #     label.image = photo  # avoid garbage collection
    #
    # image = Image.open('images.jpg')
    # copy_of_image = image.copy()
    # photo = ImageTk.PhotoImage(image)
    # label = ttk.Label(root1, image=photo)
    # label.bind('<Configure>', resize_image)
    # label.pack(fill=BOTH, expand=YES)

    title = Label(root1, text="Register Here", font=("Impact", 35, "bold"), fg="#d77337")
    title.place(x=90, y=30)

    email = Label(root1, text="Enter Email Id::", font=("bold", 10))
    email.place(x=20, y=100)

    name = Label(root1, text="Enter Name::", font=("bold", 10))
    name.place(x=20, y=130)

    uname = Label(root1, text="Enter Username::", font=("bold", 10))
    uname.place(x=20, y=160)

    password = Label(root1, text="Enter password::", font=("bold", 10))
    password.place(x=20, y=190)

    eMail = Entry(root1, font=("times new roman", 12))
    eMail.place(x=150, y=100, height=25, width=350)

    eName = Entry(root1, font=("times new roman", 12))
    eName.place(x=150, y=130, height=25, width=350)

    euName = Entry(root1, font=("times new roman", 12))
    euName.place(x=150, y=160, height=25, width=350)

    ePass = Entry(root1, font=("times new roman", 12))
    ePass.place(x=150, y=190, height=25, width=350)

    insert = Button(root1, text="Sign Up", font=("bold", 15), bg="pink", command=lambda: reg(eMail, euName, ePass,
                                                                                             eName, root1))
    insert.place(x=170, y=230)
    done = Button(root1, text="Done", font=("bold", 15), bg="pink", command=lambda: des(root1))
    done.place(x=300, y=230)
    # print(type(root1))
    root1.mainloop()

# register()